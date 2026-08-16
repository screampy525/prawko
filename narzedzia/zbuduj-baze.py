# -*- coding: utf-8 -*-
"""Zamienia oficjalny katalog XLSX na pytania.json używany przez aplikację.

Uruchomienie:  python narzedzia/zbuduj-baze.py
Wynik:         app/data/pytania.json
"""
import json
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'KATALOG_dla_kandydatów_na_kierowców_072026.xlsx')
MEDIA = os.path.join(ROOT, 'media')
OUT = os.path.join(ROOT, 'app', 'data', 'pytania.json')

import openpyxl

# Kolumny w arkuszu (0-indeksowane). Kolumny 11-14 to tłumaczenie migowe (PJM) - pomijamy.
NR, TRESC, ODP_A, ODP_B, ODP_C, POPRAWNA, MEDIA_COL, ZAKRES, PUNKTY, KATEGORIE = 1, 2, 3, 4, 5, 6, 7, 8, 9, 10


def tekst(v):
    if v is None:
        return ''
    s = str(v).replace('\r\n', '\n').strip()
    s = re.sub(r'\n{3,}', '\n\n', s)
    return s


def zbuduj_indeks_mediow():
    """Mapa: nazwa pliku z katalogu (bez rozszerzenia, małe litery) -> realna nazwa w media/."""
    idx = {}
    if not os.path.isdir(MEDIA):
        return idx
    for f in os.listdir(MEDIA):
        baza, _ = os.path.splitext(f)
        idx[baza.lower()] = f
    return idx


def main():
    media_idx = zbuduj_indeks_mediow()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    pytania = []
    braki = []
    widziane = set()

    for arkusz, w_weryfikacji in (('katalog', False), ('W trakcie weryfikacji', True)):
        if arkusz not in wb.sheetnames:
            continue
        for wiersz in wb[arkusz].iter_rows(min_row=2, values_only=True):
            nr = tekst(wiersz[NR])
            tresc = tekst(wiersz[TRESC])
            if not nr or not tresc:
                continue
            if nr in widziane:
                continue
            widziane.add(nr)

            a, b, c = (tekst(wiersz[i]) for i in (ODP_A, ODP_B, ODP_C))
            poprawna = tekst(wiersz[POPRAWNA]).upper()
            ma_abc = bool(a or b or c)
            typ = 'abc' if ma_abc else 'tn'

            # Zakres bywa zapisany różnie (literówki w źródle: "Specajlistyczny").
            zakres_raw = tekst(wiersz[ZAKRES]).upper()
            zakres = 'S' if zakres_raw.startswith('SPEC') or zakres_raw.startswith('SPECA') else 'P'

            try:
                punkty = int(float(tekst(wiersz[PUNKTY]) or 0))
            except ValueError:
                punkty = 0

            kategorie = [k.strip().upper() for k in tekst(wiersz[KATEGORIE]).split(',') if k.strip()]

            plik = tekst(wiersz[MEDIA_COL])
            media_nazwa = None
            media_typ = None
            brak = False
            if plik:
                baza, ext = os.path.splitext(plik)
                realny = media_idx.get(baza.lower())
                if realny:
                    media_nazwa = realny
                    media_typ = 'vid' if os.path.splitext(realny)[1].lower() in ('.mp4', '.webm') else 'img'
                else:
                    brak = True
                    braki.append({'nr': nr, 'plik': plik, 'arkusz': arkusz})

            # Pytanie jest "sprawne" tylko jeśli da się na nie odpowiedzieć:
            # ma poprawną odpowiedź, punktację i nie brakuje mu multimediów.
            sprawne = (
                not brak
                and punkty in (1, 2, 3)
                and bool(kategorie)
                and ((typ == 'tn' and poprawna in ('T', 'N')) or (typ == 'abc' and poprawna in ('A', 'B', 'C')))
            )

            p = {
                'id': nr,
                'tresc': tresc,
                'typ': typ,
                'poprawna': poprawna,
                'zakres': zakres,
                'punkty': punkty,
                'kategorie': kategorie,
                'media': media_nazwa,
                'mediaTyp': media_typ,
                'brakMediow': brak,
                'weryfikacja': w_weryfikacji,
                'sprawne': sprawne,
            }
            if ma_abc:
                p['odpowiedzi'] = {'A': a, 'B': b, 'C': c}
            pytania.append(p)

    # Świadomie nie wykrywamy pułapek automatycznie: to ocena, której nie da się
    # wyprowadzić z samego katalogu - pomyłka bywa zwykłą luką w wiedzy, a nie
    # dowodem, że pytanie zastawia pułapkę. Oznaczenia robi użytkownik podczas
    # nauki; te utrwalone skryptem zapisz-podchwytliwe.py wczytujemy tutaj.
    domyslne = {}
    plik_domyslnych = os.path.join(ROOT, 'narzedzia', 'podchwytliwe.json')
    if os.path.exists(plik_domyslnych):
        with open(plik_domyslnych, encoding='utf-8') as fh:
            domyslne = json.load(fh)

    for p in pytania:
        uwaga = domyslne.get(p['id'])
        p['podchwytliwe'] = {'powod': uwaga} if uwaga is not None else None

    # Stabilna kolejność: po numerze pytania.
    pytania.sort(key=lambda p: (int(p['id']) if p['id'].isdigit() else 10 ** 9, p['id']))

    kategorie_zbior = sorted({k for p in pytania for k in p['kategorie']})
    dane = {
        'wersja': '07.2026',
        'zrodlo': os.path.basename(XLSX),
        'kategorie': kategorie_zbior,
        'pytania': pytania,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as fh:
        json.dump(dane, fh, ensure_ascii=False, separators=(',', ':'))

    sprawnych = sum(1 for p in pytania if p['sprawne'])
    print(f'zapisano {len(pytania)} pytań ({sprawnych} sprawnych) -> {OUT}')
    print(f'rozmiar: {os.path.getsize(OUT) / 1048576:.2f} MB')
    print(f'brakujące multimedia: {len(braki)}')
    if braki:
        with open(os.path.join(ROOT, 'narzedzia', 'brakujace-media.json'), 'w', encoding='utf-8') as fh:
            json.dump(braki, fh, ensure_ascii=False, indent=1)

    # Podsumowanie dostępności pytań w rozbiciu na kategorie i koszyki egzaminacyjne.
    print('\nkategoria | podstawowe 3/2/1 | specjalistyczne 3/2/1')
    for k in kategorie_zbior:
        pod = [sum(1 for p in pytania if p['sprawne'] and k in p['kategorie']
                   and p['zakres'] == 'P' and p['typ'] == 'tn' and p['punkty'] == n) for n in (3, 2, 1)]
        spec = [sum(1 for p in pytania if p['sprawne'] and k in p['kategorie']
                    and p['zakres'] == 'S' and p['typ'] == 'abc' and p['punkty'] == n) for n in (3, 2, 1)]
        ok = pod[0] >= 10 and pod[1] >= 6 and pod[2] >= 4 and spec[0] >= 6 and spec[1] >= 4 and spec[2] >= 2
        print(f'{k:>4}      | {pod[0]:5} {pod[1]:4} {pod[2]:4} | {spec[0]:6} {spec[1]:4} {spec[2]:4}  {"" if ok else "<-- za mało na pełny egzamin"}')


if __name__ == '__main__':
    sys.exit(main())
